"""
Builds the multi-sheet Excel QA report:
  Spelling | Spacing | Grammar | Terminology | Duplicates | Clean

Every category check STILL runs on every string independently (nothing is
skipped) - but each string now appears in exactly ONE sheet, the single
best-fitting category, rather than being duplicated into every sheet whose
check found something.

Which sheet wins (updated 2026-08-25, per MRG's explicit instruction): the
category with the MOST individual issues for that specific string - e.g. a
string with 3 grammar problems and 1 spelling problem goes to Grammar, not
Spelling. This is a per-string count, not a fixed category ranking. Ties
(equal counts, including the common "exactly one issue in each of two
categories" case) fall back to a fixed order - Spelling > Grammar >
Terminology > Spacing - as a deterministic tie-break; MRG has not specified
a preferred tie-break, so this uses the previously-agreed default (Section 8
Q1 in CLAUDE.md). Flag if a different tie-break is wanted.
The Suggested Change column always reflects ALL fixes together regardless
of which single sheet the row landed in.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.engine import LINE_PATTERN, parse_entries
from core.language_rules import detect_language, check_entry, check_spelling_grammar_lt

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Used only as a TIE-BREAK when two or more categories have the same (nonzero)
# issue count for a string - see module docstring. Not a fixed ranking anymore.
CATEGORY_PRIORITY = ["Spelling", "Grammar", "Terminology", "Spacing"]


def _pick_category(issues):
    """Return the category with the most issues for this string, breaking
    ties with CATEGORY_PRIORITY order. Returns None if every category is
    empty (the string is Clean)."""
    counts = {c: len(issues[c]) for c in CATEGORY_PRIORITY}
    max_count = max(counts.values())
    if max_count == 0:
        return None
    return next(c for c in CATEGORY_PRIORITY if counts[c] == max_count)


def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for r in rows:
        ws.append(r)
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = max([len(str(headers[col - 1]))] + [len(str(r[col - 1])) for r in rows] + [10])
        ws.column_dimensions[letter].width = min(max_len + 2, 80)
    ws.freeze_panes = "A2"


def build_report(path, out_path, lang_override=None, lt_client=None):
    entries = parse_entries(path)
    lang_code, lang_name, confidence = (
        (lang_override, KNOWN_NAME(lang_override), "manual") if lang_override
        else detect_language(path)
    )

    from core.engine import load_glossary, fix_value
    # BUG FIX (found 2026-08-25): this used to call load_glossary() with no
    # path, which silently falls back to the 8-entry DEFAULT_GLOSSARY and
    # never reads glossary.json at all - so anything added to glossary.json
    # (e.g. "idh" -> "IDH") had ZERO effect on the actual Excel report,
    # even though README.md explicitly says "no code changes needed" to
    # extend the glossary. app.py and cli.py already did this correctly
    # (GLOSSARY_PATH = .../glossary.json passed to load_glossary); this was
    # the one place that didn't match. Verified: before this fix, a
    # glossary.json-only entry never showed up in cli_excel.py's report.
    glossary_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "glossary.json"
    )
    glossary = load_glossary(glossary_path)

    rows_by_category = {c: [] for c in CATEGORY_PRIORITY}
    duplicate_rows = []
    clean_rows = []
    seen_keys = {}

    for e in entries:
        lineno, key, value = e["line"], e["key"], e["value"]
        issues = check_entry(value, lang_code, glossary)

        # "Suggested Change" = the fully corrected string (ALL safe fixes
        # applied together, regardless of which single category the row
        # is filed under) - always review before applying, never auto-applied.
        suggested, _cats = fix_value(value, glossary, use_language_tool=False, lang=lang_code)
        if suggested == value:
            suggested = "(needs manual review - no confident auto-suggestion)"

        # Pick the single best-fitting sheet for this string - whichever
        # category has the MOST issues for THIS string (see module docstring).
        chosen_category = _pick_category(issues)

        if chosen_category == "Spelling":
            from core.spellcheck import check_spelling, get_custom_words
            hits = check_spelling(value, lang_code, get_custom_words()) or []
            words_col = ";  ".join(
                f'"{w}" -> "{s[0]}"' + (f" (or: {', '.join(s[1:3])})" if len(s) > 1 else "")
                if s else f'"{w}" - no confident suggestion'
                for w, s in hits
            )
            rows_by_category["Spelling"].append([lineno, key, value, suggested, words_col])
        elif chosen_category is not None:
            rows_by_category[chosen_category].append([lineno, key, value, suggested])

        if key in seen_keys:
            duplicate_rows.append([lineno, key, value, f"Also defined at line {seen_keys[key]}"])
        else:
            seen_keys[key] = lineno

        if chosen_category is None:
            clean_rows.append([lineno, key, value])

    wb = Workbook()
    wb.remove(wb.active)

    headers4 = ["Line", "Resource Key", "Current Value", "Suggested Change"]
    headers5 = headers4 + ["Words To Correct"]

    ws = wb.create_sheet("Spelling")
    _write_sheet(ws, headers5, rows_by_category["Spelling"])
    from core.spellcheck import is_available as spell_available
    if not spell_available(lang_code):
        ws.cell(row=1, column=7, value=f"Note: no spelling dictionary installed for '{lang_code}' - spelling not checked, sheet reflects that, not 'no errors'")

    ws = wb.create_sheet("Grammar")
    _write_sheet(ws, headers4, rows_by_category["Grammar"])

    ws = wb.create_sheet("Terminology")
    _write_sheet(ws, headers4, rows_by_category["Terminology"])

    ws = wb.create_sheet("Spacing")
    _write_sheet(ws, headers4, rows_by_category["Spacing"])

    ws = wb.create_sheet("Duplicates")
    _write_sheet(ws, ["Line", "Resource Key", "Current Value", "Note"], duplicate_rows)

    ws = wb.create_sheet("Clean")
    _write_sheet(ws, ["Line", "Resource Key", "Current Value"], clean_rows)

    ws = wb.create_sheet("Summary", 0)
    _write_sheet(ws, ["Metric", "Value"], [
        ["File", os.path.basename(path)],
        ["Detected language", f"{lang_name} ({lang_code})"],
        ["Language detection method", confidence],
        ["Total strings", len(entries)],
        ["Sheet assignment", "Each string appears in exactly ONE sheet - priority: Spelling > Grammar > Terminology > Spacing"],
        ["Spelling issues", len(rows_by_category["Spelling"])],
        ["Grammar issues", len(rows_by_category["Grammar"])],
        ["Terminology issues", len(rows_by_category["Terminology"])],
        ["Spacing issues", len(rows_by_category["Spacing"])],
        ["Duplicate keys", len(duplicate_rows)],
        ["Clean strings (no issues in any category)", len(clean_rows)],
        ["LanguageTool server used", "yes" if lt_client else "no - spelling sheet incomplete"],
    ])

    wb.save(out_path)
    return {
        "total": len(entries),
        "spelling": len(rows_by_category["Spelling"]),
        "grammar": len(rows_by_category["Grammar"]),
        "terminology": len(rows_by_category["Terminology"]),
        "spacing": len(rows_by_category["Spacing"]),
        "duplicates": len(duplicate_rows), "clean": len(clean_rows),
        "language": lang_name, "lang_code": lang_code, "detection": confidence,
    }


def KNOWN_NAME(code):
    from core.language_rules import KNOWN_LANG_CODES
    return KNOWN_LANG_CODES.get(code, code)
