"""
Builds the multi-sheet Excel QA report:
  Spelling | Spacing | Grammar | Terminology | Duplicates | Clean

Every string is run through every category check independently - a string
with issues in 2 categories gets a row in both sheets. A string only lands
in "Clean" if every category came back empty.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.engine import LINE_PATTERN, parse_entries
from core.language_rules import detect_language, check_entry, check_spelling_grammar_lt

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for r in rows:
        ws.append(r)
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = max([len(str(headers[col - 1]))] + [len(str(r[col - 1])) for r in rows] + [10])
        ws.column_dimensions[letter].width = min(max_len + 2, 80)
    ws.freeze_panes = "A2"


def build_report(path, out_path, lang_override=None, lt_client=None):
    entries = parse_entries(path)
    lang_code, lang_name, confidence = (
        (lang_override, KNOWN_NAME(lang_override), "manual") if lang_override
        else detect_language(path)
    )

    spelling_rows, spacing_rows, grammar_rows, terminology_rows = [], [], [], []
    duplicate_rows = []
    clean_rows = []
    seen_keys = {}

    for e in entries:
        lineno, key, value = e["line"], e["key"], e["value"]
        issues = check_entry(value, lang_code)

        lt_spelling, lt_grammar = check_spelling_grammar_lt(value, lang_code, lt_client)
        issues["Spelling"].extend(lt_spelling)
        issues["Grammar"].extend(lt_grammar)

        for issue in issues["Spelling"]:
            spelling_rows.append([lineno, key, value, issue])
        for issue in issues["Spacing"]:
            spacing_rows.append([lineno, key, value, issue])
        for issue in issues["Grammar"]:
            grammar_rows.append([lineno, key, value, issue])
        for issue in issues["Terminology"]:
            terminology_rows.append([lineno, key, value, issue])

        if key in seen_keys:
            duplicate_rows.append([lineno, key, value, f"Also defined at line {seen_keys[key]}"])
        else:
            seen_keys[key] = lineno

        if not any(issues.values()):
            clean_rows.append([lineno, key, value])

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Spelling")
    _write_sheet(ws, ["Line", "Key", "String", "Issue"], spelling_rows)
    if lt_client is None:
        ws.cell(row=1, column=6, value="Note: no LanguageTool server configured - spelling not checked, sheet reflects that, not 'no errors'")

    ws = wb.create_sheet("Spacing")
    _write_sheet(ws, ["Line", "Key", "String", "Issue"], spacing_rows)

    ws = wb.create_sheet("Grammar")
    _write_sheet(ws, ["Line", "Key", "String", "Issue"], grammar_rows)

    ws = wb.create_sheet("Terminology")
    _write_sheet(ws, ["Line", "Key", "String", "Issue"], terminology_rows)

    ws = wb.create_sheet("Duplicates")
    _write_sheet(ws, ["Line", "Key", "String", "Note"], duplicate_rows)

    ws = wb.create_sheet("Clean")
    _write_sheet(ws, ["Line", "Key", "String"], clean_rows)

    ws = wb.create_sheet("Summary", 0)
    _write_sheet(ws, ["Metric", "Value"], [
        ["File", os.path.basename(path)],
        ["Detected language", f"{lang_name} ({lang_code})"],
        ["Language detection method", confidence],
        ["Total strings", len(entries)],
        ["Spelling issues", len(spelling_rows)],
        ["Spacing issues", len(spacing_rows)],
        ["Grammar issues", len(grammar_rows)],
        ["Terminology issues", len(terminology_rows)],
        ["Duplicate keys", len(duplicate_rows)],
        ["Clean strings (no issues in any category)", len(clean_rows)],
        ["LanguageTool server used", "yes" if lt_client else "no - spelling sheet incomplete"],
    ])

    wb.save(out_path)
    return {
        "total": len(entries), "spelling": len(spelling_rows), "spacing": len(spacing_rows),
        "grammar": len(grammar_rows), "terminology": len(terminology_rows),
        "duplicates": len(duplicate_rows), "clean": len(clean_rows),
        "language": lang_name, "lang_code": lang_code, "detection": confidence,
    }


def KNOWN_NAME(code):
    from core.language_rules import KNOWN_LANG_CODES
    return KNOWN_LANG_CODES.get(code, code)
